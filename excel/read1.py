import openpyxl
wb=openpyxl.load_workbook('example.xlsx')
print(wb)
print(wb.get_sheet_names())#to get diff sheet names in workbook
print(wb.active)#to get active sheet
a=wb.get_sheet_by_name('Sheet1')#to get sheet by name
print(a)
print(a.title)#to get title of a
print(a['A1'])#to get cells
print(a['A1'].value)


a.cell(row=1, column=2)
a.cell(row=1, column=2).value
for i in range(1, 8, 2):
    print(i,   a.cell(row=i, column=2).value)

print(a.max_row)
print(a.max_column)
