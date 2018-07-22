import openpyxl
wb = openpyxl.Workbook()
print(wb.get_sheet_names())
wb.create_sheet()
print(wb.get_sheet_names())
wb.create_sheet(index=0, title='First Sheet')
print(wb.get_sheet_names())
wb.create_sheet(index=2, title='Middle Sheet')
print(wb.get_sheet_names())
"""
The create_sheet() method returns a new Worksheet object named
SheetX, which by default is set to be the last sheet in the workbook.
Optionally, the index and name of the new sheet can be specified with
the index and title keyword arguments.
"""

wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
wb.remove_sheet(wb.get_sheet_by_name('Sheet1'))
print(wb.get_sheet_names())
wb.save('j2.xlsx')